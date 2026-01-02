from rest_framework import generics, viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.exceptions import ValidationError, PermissionDenied
from django.shortcuts import get_object_or_404
from django.db.models import Count, Sum, Q
from django.db import transaction
from django.utils.timezone import now, timedelta, make_aware
from django.core.management.base import CommandError
from django.core.files.base import ContentFile
from django.conf import settings
import qrcode
import io, csv
import openpyxl
from datetime import datetime, time, timedelta
import secrets, string
import threading

from .models import (
    User, Student, Mentor, ClubOrganizer, Club, Event, EventRegistration, Certificate,
    Hall, HallBooking, AICTECategory, AICTEPointTransaction, Notification, AuditLog,
    ClubMember, ClubRole, EventAttendance, CertificateTemplate, PrincipalSignature
)
from .serializers import (
    UserSerializer, RegisterSerializer, EventSerializer,
    CertificateSerializer, EventRegistrationSerializer, ClubSerializer,
    HallSerializer, HallBookingSerializer, AICTECategorySerializer, AICTEPointTransactionSerializer,
    NotificationSerializer, AuditLogSerializer, StudentSerializer, StudentProfileSerializer,
    MentorSerializer, MentorProfileSerializer, ClubOrganizerSerializer, ClubOrganizerProfileSerializer,
    ClubMemberSerializer, ClubRoleSerializer, EventAttendanceSerializer, CertificateTemplateSerializer,
    PrincipalSignatureSerializer
)
from .permissions import IsClubAdmin, IsStudent, IsMentor, IsAdmin
from .email_utils import send_verification_email, send_password_reset_email, send_account_locked_email
from .certificate_generator import CertificateGenerator


def log_action(user, action):
    """Log user actions for audit trail"""
    AuditLog.objects.create(user=user, action=action)


# ============================================================================
# AUTHENTICATION & REGISTRATION
# ============================================================================

class RegisterView(generics.CreateAPIView):
    """
from .certificate_generator import CertificateGenerator
    User registration endpoint supporting students, mentors, and club organizers.
    Requires email verification before account activation.
    """
    queryset = User.objects.all()
    permission_classes = (AllowAny,)
    serializer_class = RegisterSerializer

    def perform_create(self, serializer):
        user = serializer.save()
        # Send verification email
        try:
            threading.Thread(target=send_verification_email, args=(user,)).start()
        except Exception as e:
            print(f"Error sending verification email: {e}")
        log_action(user, f"User registered: {user.username}")

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        response.data['message'] = "Registration successful! Please check your email to verify your account."
        return response


class VerifyEmailView(generics.GenericAPIView):
    """
    Email verification endpoint. Marks user as verified when valid token is provided.
    """
    permission_classes = (AllowAny,)
    
    def post(self, request):
        token = request.data.get('token')
        if not token:
            return Response(
                {'error': 'Verification token is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = User.objects.filter(email_verification_token=token).first()
        if not user:
            return Response(
                {'error': 'Invalid or expired verification token.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.is_email_verified = True
        user.email_verification_token = None
        user.save()
        log_action(user, "Email verified")
        
        return Response(
            {'message': 'Email verified successfully! You can now log in.'},
            status=status.HTTP_200_OK
        )


class ResendVerificationEmailView(generics.GenericAPIView):
    """
    Resend verification email to user.
    """
    permission_classes = (AllowAny,)
    
    def post(self, request):
        email = request.data.get('email')
        if not email:
            return Response(
                {'error': 'Email is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = User.objects.filter(email=email).first()
        if not user:
            return Response(
                {'error': 'User with this email not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if user.is_email_verified:
            return Response(
                {'error': 'Email is already verified.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate new verification token
        import uuid
        user.email_verification_token = str(uuid.uuid4())
        user.verification_sent_at = now()
        user.save()
        
        try:
            threading.Thread(target=send_verification_email, args=(user,)).start()
        except Exception as e:
            print(f"Error sending verification email: {e}")
        log_action(user, "Verification email resent")
        
        return Response(
            {'message': 'Verification email resent successfully!'},
            status=status.HTTP_200_OK
        )


class RequestPasswordResetView(generics.GenericAPIView):
    """
    Request password reset OTP (6 digits, 10-minute validity).
    """
    permission_classes = (AllowAny,)
    
    def post(self, request):
        email = request.data.get('email')
        if not email:
            return Response(
                {'error': 'Email is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = User.objects.filter(email=email).first()
        if not user:
            return Response(
                {'error': 'User with this email not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Generate 6-digit OTP
        otp = ''.join(secrets.choice(string.digits) for _ in range(6))
        user.password_reset_token = otp
        user.password_reset_expires = now() + timedelta(minutes=10)
        user.save()
        
        try:
            threading.Thread(target=send_password_reset_email, args=(user, otp)).start()
        except Exception as e:
            print(f"Error sending password reset email: {e}")
        log_action(user, "Password reset requested")
        
        return Response(
            {'message': 'Password reset OTP sent to your email.'},
            status=status.HTTP_200_OK
        )


class ResetPasswordView(generics.GenericAPIView):
    """
    Reset password using OTP and new password.
    """
    permission_classes = (AllowAny,)
    
    def post(self, request):
        email = request.data.get('email')
        otp = request.data.get('otp')
        new_password = request.data.get('new_password')
        
        if not all([email, otp, new_password]):
            return Response(
                {'error': 'Email, OTP, and new password are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = User.objects.filter(email=email).first()
        if not user:
            return Response(
                {'error': 'User not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if not user.password_reset_token or user.password_reset_token != otp:
            return Response(
                {'error': 'Invalid OTP.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if user.password_reset_expires < now():
            return Response(
                {'error': 'OTP has expired.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.set_password(new_password)
        user.password_reset_token = None
        user.password_reset_expires = None
        user.save()
        log_action(user, "Password reset successfully")
        
        return Response(
            {'message': 'Password reset successfully!'},
            status=status.HTTP_200_OK
        )


class LoginView(generics.GenericAPIView):
    """
    Custom login view handling failed login attempts and account lockout.
    Requires email verification before allowing login.
    """
    permission_classes = (AllowAny,)
    
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        
        if not username or not password:
            return Response(
                {'error': 'Username and password are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = User.objects.filter(Q(username=username) | Q(email=username)).first()
        
        if not user:
            return Response(
                {'error': 'Invalid credentials.'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        # Check if account is locked
        if user.account_locked_until and user.account_locked_until > now():
            return Response(
                {'error': 'Account is locked. Try again later.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Check email verification (skip for admin users)
        if not user.is_email_verified and user.user_type != 'admin':
            return Response(
                {'error': 'Please verify your email before logging in.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Verify password
        if not user.check_password(password):
            user.failed_login_attempts += 1
            
            # Lock account after 5 failed attempts
            if user.failed_login_attempts >= 5:
                user.account_locked_until = now() + timedelta(minutes=15)
            
            user.save()
            log_action(user, f"Failed login attempt ({user.failed_login_attempts})")
            
            return Response(
                {'error': 'Invalid credentials.'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        # Successful login - reset failed attempts
        user.failed_login_attempts = 0
        user.account_locked_until = None
        user.save()
        log_action(user, "User logged in")
        
        from rest_framework_simplejwt.tokens import RefreshToken
        
        refresh = RefreshToken.for_user(user)
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'user': UserSerializer(user).data
        }, status=status.HTTP_200_OK)


# ============================================================================
# PROFILE MANAGEMENT
# ============================================================================

class ProfileView(generics.RetrieveUpdateAPIView):
    """
    Returns role-specific profile. Updates only username, first_name, last_name.
    """
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        user = self.request.user
        if user.user_type == 'student' and hasattr(user, 'student_profile'):
            return StudentProfileSerializer
        elif user.user_type == 'mentor' and hasattr(user, 'mentor_profile'):
            return MentorProfileSerializer
        elif user.user_type == 'club_organizer' and hasattr(user, 'club_organizer_profile'):
            return ClubOrganizerProfileSerializer
        return UserSerializer  # fallback, unlikely

    def get_object(self):
        user = self.request.user
        if user.user_type == 'student' and hasattr(user, 'student_profile'):
            return user.student_profile
        elif user.user_type == 'mentor' and hasattr(user, 'mentor_profile'):
            return user.mentor_profile
        elif user.user_type == 'club_organizer' and hasattr(user, 'club_organizer_profile'):
            return user.club_organizer_profile
        return user

    def perform_update(self, serializer):
        """
        Override perform_update to add logging and ensure user model updates are saved
        """
        instance = self.get_object()

        # Extract user data and profile data
        user_data = {}
        profile_data = {}

        for key, value in serializer.validated_data.items():
            if key.startswith('user.'):
                user_data[key[5:]] = value  # Remove 'user.' prefix
            else:
                profile_data[key] = value
                
        # Update user model if user data was provided
        if user_data:
            user = getattr(instance, 'user', instance)
            for attr, value in user_data.items():
                setattr(user, attr, value)
            user.save()

        # Update profile model
        for attr, value in profile_data.items():
            setattr(instance, attr, value)
        instance.save()

        # Check profile completion for profiles
        if hasattr(instance, 'profile_completed'):
            required_fields = []
            if hasattr(instance, 'student_profile'):
                required_fields = ['phone_number', 'date_of_birth', 'address', 'emergency_contact_name', 'emergency_contact_phone']
            elif hasattr(instance, 'mentor_profile'):
                required_fields = ['phone_number', 'date_of_birth', 'address', 'qualifications']
            elif hasattr(instance, 'club_organizer_profile'):
                required_fields = ['phone_number', 'date_of_birth', 'address', 'designation_in_club']

            if required_fields and not instance.profile_completed:
                if all(getattr(instance, field, None) for field in required_fields):
                    instance.profile_completed = True
                    instance.profile_completed_at = now()
                    instance.save()

        log_action(self.request.user, f"Updated profile ({self.request.user.user_type})")




class StudentProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = StudentProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return get_object_or_404(Student, user=self.request.user)

    def perform_update(self, serializer):
        student = serializer.save()
        user_data = self.request.data.get('user', {})
        user = self.request.user
        for attr in ['username', 'first_name', 'last_name']:
            if attr in user_data:
                setattr(user, attr, user_data[attr])
        user.save()

        required_fields = [
            'phone_number', 'date_of_birth', 'address',
            'emergency_contact_name', 'emergency_contact_phone'
        ]
        if all(getattr(student, field, None) for field in required_fields) and not student.profile_completed:
            student.profile_completed = True
            student.profile_completed_at = now()
            student.save()

        log_action(user, "Updated student profile")
    



class MentorProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = MentorProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return get_object_or_404(Mentor, user=self.request.user)

    def perform_update(self, serializer):
        mentor = serializer.save()
        user_data = self.request.data.get('user', {})
        user = self.request.user
        for attr in ['username', 'first_name', 'last_name']:
            if attr in user_data:
                setattr(user, attr, user_data[attr])
        user.save()

        required_fields = [
            'phone_number', 'date_of_birth', 'address', 'qualifications'
        ]
        if all(getattr(mentor, field, None) for field in required_fields) and not mentor.profile_completed:
            mentor.profile_completed = True
            mentor.profile_completed_at = now()
            mentor.save()

        log_action(user, "Updated mentor profile")


class ClubOrganizerProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = ClubOrganizerProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return get_object_or_404(ClubOrganizer, user=self.request.user)

    def perform_update(self, serializer):
        organizer = serializer.save()
        user_data = self.request.data.get('user', {})
        user = self.request.user
        for attr in ['username', 'first_name', 'last_name']:
            if attr in user_data:
                setattr(user, attr, user_data[attr])
        user.save()

        required_fields = [
            'phone_number', 'date_of_birth', 'address', 'designation_in_club'
        ]
        if all(getattr(organizer, field, None) for field in required_fields) and not organizer.profile_completed:
            organizer.profile_completed = True
            organizer.profile_completed_at = now()
            organizer.save()

        log_action(user, "Updated club organizer profile")


# ============================================================================
# STUDENT MANAGEMENT (for mentee assignment and dashboard)
# ============================================================================

class StudentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Student list and details endpoint.
    - GET /api/students/ -> List all students (admin/mentor can see all)
    - GET /api/students/mentees -> List students assigned to current mentor
    """
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        # Admin can see all students
        if user.user_type == 'admin':
            return Student.objects.all()
        # Mentor can only see their mentees
        elif user.user_type == 'mentor':
            mentor = getattr(user, 'mentor_profile', None)
            if mentor:
                return Student.objects.filter(mentor=mentor)
            return Student.objects.none()
        # Students can only see themselves (not really useful but included)
        elif user.user_type == 'student':
            return Student.objects.filter(user=user)
        return Student.objects.none()
    
    @action(detail=False, methods=['get'])
    def mentees(self, request):
        """
        List students assigned to the current mentor.
        GET /api/students/mentees
        """
        if request.user.user_type != 'mentor':
            return Response(
                {'detail': 'Only mentors can view their mentees.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        mentor = getattr(request.user, 'mentor_profile', None)
        if not mentor:
            return Response(
                {'detail': 'Mentor profile not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        mentees = Student.objects.filter(mentor=mentor)
        serializer = self.get_serializer(mentees, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        """
        Assign a student to the current mentor.
        POST /api/students/{id}/assign/
        """
        if request.user.user_type != 'mentor':
            return Response(
                {'detail': 'Only mentors can assign mentees.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        student = self.get_object()
        mentor = getattr(request.user, 'mentor_profile', None)
        
        if not mentor:
            return Response(
                {'detail': 'Mentor profile not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        student.mentor = mentor
        student.save()
        log_action(request.user, f"Assigned student {student.user.username} as mentee")
        
        return Response(
            {'message': f'Student {student.user.username} assigned successfully.'},
            status=status.HTTP_200_OK
        )
        
# -----------------------
# Mentor Endpoints
# -----------------------
class MentorViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Mentor list and detail endpoints.
    - GET /api/mentors/ -> list all mentors (admin only)
    - GET /api/mentors/{id}/mentees/ -> list mentees of a mentor
    """
    queryset = Mentor.objects.all()
    serializer_class = MentorSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'admin':
            return Mentor.objects.all()
        return Mentor.objects.none()

    @action(detail=True, methods=['get'])
    def mentees(self, request, pk=None):
        """List students assigned to this mentor"""
        mentor = self.get_object()
        mentees = Student.objects.filter(mentor=mentor)
        from .serializers import StudentSerializer
        serializer = StudentSerializer(mentees, many=True)
        return Response(serializer.data)

# -----------------------
# Club Organizer Endpoints
# -----------------------
class ClubOrganizerViewSet(viewsets.ModelViewSet):
    """
    Club organizer endpoints.
    - GET /api/club-organizers/ -> list all organizers (admin only)
    - POST /api/club-organizers/{id}/assign-club/ -> assign organizer to a club
    """
    queryset = ClubOrganizer.objects.all()
    serializer_class = ClubOrganizerSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'admin':
            return ClubOrganizer.objects.all()
        return ClubOrganizer.objects.none()

    @action(detail=True, methods=['post'])
    def assign_club(self, request, pk=None):
        """Assign this organizer to a club"""
        club_id = request.data.get('club_id')
        if not club_id:
            return Response({'error': 'club_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        organizer = self.get_object()
        club = get_object_or_404(Club, id=club_id)

        organizer.club = club
        organizer.save()
        log_action(request.user, f"Assigned {organizer.user.username} to club {club.name}")
        return Response({'message': f'{organizer.user.get_full_name()} assigned to {club.name}'}, status=status.HTTP_200_OK)


# ============================================================================
# ADMIN USER MANAGEMENT
# ============================================================================

class AdminUserCreationView(generics.CreateAPIView):
    """
    Admin endpoint for creating users (students, mentors, club organizers).
    Supports bulk user creation via CSV.
    """
    queryset = User.objects.all()
    permission_classes = [IsAuthenticated]
    serializer_class = RegisterSerializer
    
    def create(self, request, *args, **kwargs):
        # Check admin permission
        if request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can create users.")
        
        return super().create(request, *args, **kwargs)
    
    def perform_create(self, serializer):
        user = serializer.save()
        # Auto-verify email for admin-created accounts
        user.is_email_verified = True
        user.email_verification_token = None
        user.save()
        log_action(self.request.user, f"Created user account: {user.username}")


class BulkUserCreationView(generics.GenericAPIView):
    """
    Bulk user creation from CSV file.
    CSV format: username,email,first_name,last_name,user_type,usn/employee_id,department,semester/designation
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        if request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can create users.")
        
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            return Response(
                {'error': 'CSV file is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            csv_reader = csv.DictReader(decoded_file)
            
            created_users = []
            errors = []
            
            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    username = row.get('username', '').strip()
                    email = row.get('email', '').strip()
                    first_name = row.get('first_name', '').strip()
                    last_name = row.get('last_name', '').strip()
                    user_type = row.get('user_type', '').strip()
                    department = row.get('department', '').strip()
                    
                    if not all([username, email, user_type]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        continue
                    
                    # Generate temporary password
                    temp_password = secrets.token_urlsafe(12)
                    
                    user = User.objects.create(
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        user_type=user_type,
                        is_email_verified=True
                    )
                    user.set_password(temp_password)
                    user.save()
                    
                    # Create role-specific profile
                    if user_type == 'student':
                        usn = row.get('usn', '').strip()
                        semester = int(row.get('semester', 1))
                        if not usn:
                            errors.append(f"Row {row_num}: USN required for student")
                            user.delete()
                            continue
                        Student.objects.create(
                            user=user, usn=usn, department=department, semester=semester
                        )
                    
                    elif user_type == 'mentor':
                        employee_id = row.get('employee_id', '').strip()
                        designation = row.get('designation', '').strip()
                        if not all([employee_id, designation]):
                            errors.append(f"Row {row_num}: Employee ID and Designation required for mentor")
                            user.delete()
                            continue
                        Mentor.objects.create(
                            user=user, employee_id=employee_id, department=department, designation=designation
                        )
                    
                    created_users.append({
                        'username': username,
                        'email': email,
                        'temporary_password': temp_password
                    })
                    log_action(request.user, f"Created user via bulk import: {username}")
                
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
            return Response({
                'created_count': len(created_users),
                'created_users': created_users,
                'errors': errors
            }, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response(
                {'error': f"Error processing CSV: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )


# ============================================================================
# MENTOR-MENTEE MANAGEMENT
# ============================================================================

class MentorMenteeAssignmentView(generics.GenericAPIView):
    """
    Admin endpoint for assigning/reassigning mentees to mentors.
    Supports bulk assignment via JSON.
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        if request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can manage mentor assignments.")
        
        mentor_id = request.data.get('mentor_id')
        student_ids = request.data.get('student_ids', [])
        
        if not mentor_id:
            return Response(
                {'error': 'Mentor ID is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        mentor = get_object_or_404(Mentor, id=mentor_id)
        
        for student_id in student_ids:
            try:
                student = Student.objects.get(id=student_id)
                old_mentor = student.mentor
                student.mentor = mentor
                student.save()
                log_action(request.user, f"Assigned student {student.usn} to mentor {mentor.employee_id}")
            except Student.DoesNotExist:
                pass
        
        return Response({
            'message': f'Successfully assigned {len(student_ids)} students to mentor {mentor.user.get_full_name()}',
            'assigned_count': len(student_ids)
        }, status=status.HTTP_200_OK)


class BulkMenteeAssignmentView(generics.GenericAPIView):
    """
    Bulk mentee assignment from CSV file.
    CSV format: mentor_employee_id,student_usn
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        if request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can manage mentor assignments.")
        
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            return Response(
                {'error': 'CSV file is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            csv_reader = csv.DictReader(decoded_file)
            
            assigned_count = 0
            errors = []
            
            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    mentor_emp_id = row.get('mentor_employee_id', '').strip()
                    student_usn = row.get('student_usn', '').strip()
                    
                    if not all([mentor_emp_id, student_usn]):
                        errors.append(f"Row {row_num}: Missing mentor_employee_id or student_usn")
                        continue
                    
                    mentor = Mentor.objects.filter(employee_id=mentor_emp_id).first()
                    student = Student.objects.filter(usn=student_usn).first()
                    
                    if not mentor:
                        errors.append(f"Row {row_num}: Mentor with ID {mentor_emp_id} not found")
                        continue
                    
                    if not student:
                        errors.append(f"Row {row_num}: Student with USN {student_usn} not found")
                        continue
                    
                    student.mentor = mentor
                    student.save()
                    assigned_count += 1
                    log_action(request.user, f"Bulk assigned student {student_usn} to mentor {mentor_emp_id}")
                
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
            return Response({
                'assigned_count': assigned_count,
                'errors': errors
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response(
                {'error': f"Error processing CSV: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )


# ============================================================================
# ENHANCED ADMIN MANAGEMENT FEATURES
# ============================================================================

class AdminUserListViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Admin user listing and filtering.
    Supports filtering by user_type, is_active, search by username/email.
    """
    permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAdmin]
    
    def get_queryset(self):
        queryset = User.objects.all()
        
        # Filter by user type
        user_type = self.request.query_params.get('user_type')
        if user_type:
            queryset = queryset.filter(user_type=user_type)
        
        # Filter by email verification status
        is_verified = self.request.query_params.get('is_verified')
        if is_verified:
            queryset = queryset.filter(is_email_verified=is_verified.lower() == 'true')
        
        # Search by username or email
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(Q(username__icontains=search) | Q(email__icontains=search))
        
        return queryset.order_by('-date_joined')
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def disable_account(self, request, pk=None):
        """Disable a user account"""
        user = self.get_object()
        user.is_active = False
        user.save()
        log_action(request.user, f"Disabled user account: {user.username}")
        return Response({'message': f'User {user.username} account disabled.'}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def enable_account(self, request, pk=None):
        """Enable a user account"""
        user = self.get_object()
        user.is_active = True
        user.save()
        log_action(request.user, f"Enabled user account: {user.username}")
        return Response({'message': f'User {user.username} account enabled.'}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def reset_password(self, request, pk=None):
        """Admin can manually reset user password"""
        user = self.get_object()
        new_password = request.data.get('new_password')
        
        if not new_password or len(new_password) < 8:
            return Response(
                {'error': 'Password must be at least 8 characters long.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.set_password(new_password)
        user.failed_login_attempts = 0
        user.account_locked_until = None
        user.save()
        log_action(request.user, f"Reset password for user: {user.username}")
        return Response({'message': f'Password reset for user {user.username}.'}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def unlock_account(self, request, pk=None):
        """Unlock a locked user account"""
        user = self.get_object()
        user.failed_login_attempts = 0
        user.account_locked_until = None
        user.save()
        log_action(request.user, f"Unlocked account: {user.username}")
        return Response({'message': f'User {user.username} account unlocked.'}, status=status.HTTP_200_OK)


class AdminClubManagementViewSet(viewsets.ModelViewSet):
    """
    Admin club management - ONLY admins can create clubs and assign organizers.
    """
    queryset = Club.objects.all()
    serializer_class = ClubSerializer
    permission_classes = [IsAdmin]
    
    def get_queryset(self):
        search = self.request.query_params.get('search')
        queryset = Club.objects.prefetch_related('faculty_coordinator', 'club_head', 'members')
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(description__icontains=search))
        return queryset
    
    def perform_create(self, serializer):
        """Create club with faculty coordinator."""
        coordinator_id = self.request.data.get('faculty_coordinator')
        if not coordinator_id:
            raise ValidationError("Faculty coordinator (mentor) is required")
        
        try:
            mentor = Mentor.objects.get(id=coordinator_id)
        except Mentor.DoesNotExist:
            raise ValidationError("Mentor not found")
        
        club = serializer.save(faculty_coordinator=mentor)
        log_action(self.request.user, f"Created club: {club.name} with coordinator: {mentor.user.get_full_name()}")
    
    def perform_update(self, serializer):
        club = serializer.save()
        log_action(self.request.user, f"Updated club: {club.name}")
    
    def perform_destroy(self, instance):
        club_name = instance.name
        instance.delete()
        log_action(self.request.user, f"Deleted club: {club_name}")
    
    @action(detail=True, methods=['post'])
    def assign_organizer(self, request, pk=None):
        """Assign a club organizer to this club."""
        club = self.get_object()
        organizer_id = request.data.get('organizer_id')
        
        if not organizer_id:
            return Response({'error': 'organizer_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            organizer_profile = ClubOrganizer.objects.get(id=organizer_id)
            organizer_profile.club = club
            organizer_profile.save()
            log_action(request.user, f"Assigned {organizer_profile.user.username} to club {club.name}")
            return Response(
                {'message': f'{organizer_profile.user.get_full_name()} assigned to club {club.name}'},
                status=status.HTTP_200_OK
            )
        except ClubOrganizer.DoesNotExist:
            return Response({'error': 'Club organizer not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def assign_coordinator(self, request, pk=None):
        """Change club faculty coordinator."""
        club = self.get_object()
        mentor_id = request.data.get('mentor_id')
        
        if not mentor_id:
            return Response({'error': 'mentor_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            mentor = Mentor.objects.get(id=mentor_id)
            club.faculty_coordinator = mentor
            club.save()
            log_action(request.user, f"Changed coordinator of {club.name} to {mentor.user.get_full_name()}")
            return Response(
                {'message': f'{mentor.user.get_full_name()} is now coordinator of {club.name}'},
                status=status.HTTP_200_OK
            )
        except Mentor.DoesNotExist:
            return Response({'error': 'Mentor not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def assign_club_head(self, request, pk=None):
        """Assign club head (student) to club"""
        club = self.get_object()
        student_id = request.data.get('student_id')

        if not student_id:
            return Response({'error': 'student_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.get(id=student_id)
            club.club_head = student
            club.save()
            log_action(request.user, f"Assigned {student.user.username} as club head to club {club.name}")
            return Response({'message': f'{student.user.get_full_name()} assigned as club head.'}, status=status.HTTP_200_OK)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def principal_signature(self, request):
        """Get current active principal signature"""
        signature = PrincipalSignature.objects.filter(is_active=True).first()
        if signature:
            serializer = PrincipalSignatureSerializer(signature)
            return Response(serializer.data)
        return Response({'detail': 'No active principal signature found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def upload_principal_signature(self, request):
        """Upload new principal signature"""
        if not request.FILES.get('signature_image'):
            return Response({'error': 'signature_image file is required'}, status=status.HTTP_400_BAD_REQUEST)

        notes = request.data.get('notes', '')

        # Create new signature record
        principal_signature = PrincipalSignature.objects.create(
            signature_image=request.FILES['signature_image'],
            uploaded_by=request.user,
            notes=notes
        )

        log_action(request.user, f"Uploaded new principal signature: {principal_signature}")

        serializer = PrincipalSignatureSerializer(principal_signature)
        return Response({
            'message': 'Principal signature uploaded successfully',
            'signature': serializer.data
        }, status=status.HTTP_201_CREATED)


class AdminMenteeAssignmentViewSet(viewsets.ViewSet):
    """
    Admin mentee assignment management with bulk operations.
    """
    permission_classes = [IsAdmin]
    
    def list(self, request):
        """Get all mentor-mentee assignments"""
        mentors = Mentor.objects.prefetch_related('mentees').all()
        data = []
        for mentor in mentors:
            data.append({
                'mentor_id': mentor.id,
                'mentor_name': mentor.user.get_full_name(),
                'mentor_employee_id': mentor.employee_id,
                'mentee_count': mentor.mentees.count(),
                'mentees': [{'id': s.id, 'usn': s.usn, 'name': s.user.get_full_name()} for s in mentor.mentees.all()]
            })
        return Response(data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'])
    def bulk_assign(self, request):
        """Bulk assign mentees to mentors from CSV
        CSV format: mentor_employee_id,student_usn
        """
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            return Response({'error': 'CSV file required.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            csv_reader = csv.DictReader(decoded_file)
            
            assigned = 0
            errors = []
            
            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    mentor_emp_id = row.get('mentor_employee_id', '').strip()
                    student_usn = row.get('student_usn', '').strip()
                    
                    if not all([mentor_emp_id, student_usn]):
                        errors.append(f"Row {row_num}: Missing mentor_employee_id or student_usn")
                        continue
                    
                    mentor = Mentor.objects.filter(employee_id=mentor_emp_id).first()
                    if not mentor:
                        errors.append(f"Row {row_num}: Mentor with employee ID {mentor_emp_id} not found")
                        continue
                    
                    student = Student.objects.filter(usn=student_usn).first()
                    if not student:
                        errors.append(f"Row {row_num}: Student with USN {student_usn} not found")
                        continue
                    
                    student.mentor = mentor
                    student.save()
                    assigned += 1
                    log_action(request.user, f"Assigned mentee {student_usn} to mentor {mentor_emp_id}")
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
            return Response({'assigned': assigned, 'errors': errors}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class AdminAICTEConfigViewSet(viewsets.ModelViewSet):
    """
    Admin AICTE category and rule management.
    """
    permission_classes = [IsAdmin]
    serializer_class = AICTECategorySerializer
    queryset = AICTECategory.objects.all()
    
    def perform_create(self, serializer):
        category = serializer.save()
        log_action(self.request.user, f"Created AICTE category: {category.name}")
    
    def perform_update(self, serializer):
        category = serializer.save()
        log_action(self.request.user, f"Updated AICTE category: {category.name}")
    
    def perform_destroy(self, instance):
        category_name = instance.name
        instance.delete()
        log_action(self.request.user, f"Deleted AICTE category: {category_name}")


class AdminReportingViewSet(viewsets.ViewSet):
    """
    Admin system reporting and analytics.
    """
    permission_classes = [IsAdmin]
    
    @action(detail=False, methods=['get'])
    def system_stats(self, request):
        """Get system-wide statistics"""
        stats = {
            'total_users': User.objects.count(),
            'total_students': Student.objects.count(),
            'total_mentors': Mentor.objects.count(),
            'total_clubs': Club.objects.count(),
            'total_events': Event.objects.count(),
            'total_certificates_issued': Certificate.objects.count(),
            'active_bookings': HallBooking.objects.filter(booking_status='APPROVED').count(),
            'user_breakdown': {
                'students': User.objects.filter(user_type='student').count(),
                'mentors': User.objects.filter(user_type='mentor').count(),
                'club_organizers': User.objects.filter(user_type='club_organizer').count(),
                'admins': User.objects.filter(user_type='admin').count(),
            }
        }
        return Response(stats, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def audit_logs(self, request):
        """Get filtered audit logs"""
        user_id = request.query_params.get('user_id')
        action_search = request.query_params.get('action')
        days = int(request.query_params.get('days', 30))
        
        cutoff_date = now() - timedelta(days=days)
        logs = AuditLog.objects.filter(timestamp__gte=cutoff_date).order_by('-timestamp')
        
        if user_id:
            logs = logs.filter(user_id=user_id)
        if action_search:
            logs = logs.filter(action__icontains=action_search)
        
        serializer = AuditLogSerializer(logs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def user_activity_report(self, request):
        """Get user activity statistics"""
        report = {
            'verified_users': User.objects.filter(is_email_verified=True).count(),
            'unverified_users': User.objects.filter(is_email_verified=False).count(),
            'active_users': User.objects.filter(is_active=True).count(),
            'inactive_users': User.objects.filter(is_active=False).count(),
            'locked_users': User.objects.filter(account_locked_until__isnull=False, account_locked_until__gt=now()).count(),
        }
        return Response(report, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def event_statistics(self, request):
        """Get event statistics and attendance"""
        stats = {
            'total_events': Event.objects.count(),
            'events_by_status': dict(Event.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')),
            'total_registrations': EventRegistration.objects.count(),
            'attendance_rate': {
                'attended': EventAttendance.objects.filter(is_present=True).count(),
                'absent': EventAttendance.objects.filter(is_present=False).count(),
            }
        }
        return Response(stats, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def hall_utilization_report(self, request):
        """Get hall booking statistics"""
        halls = Hall.objects.annotate(
            total_bookings=Count('bookings'),
            approved_bookings=Count('bookings', filter=Q(bookings__booking_status='APPROVED')),
            pending_bookings=Count('bookings', filter=Q(bookings__booking_status='PENDING'))
        ).values('id', 'name', 'capacity', 'total_bookings', 'approved_bookings', 'pending_bookings')
        
        return Response(list(halls), status=status.HTTP_200_OK)


# ============================================================================
# CLUB MANAGEMENT
# ============================================================================

class ClubViewSet(viewsets.ModelViewSet):
    """
    CRUD operations for clubs.
    Only admins and mentors (faculty coordinators) can create clubs.
    """
    queryset = Club.objects.all()
    serializer_class = ClubSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        if self.request.user.user_type not in ['admin', 'mentor']:
            raise PermissionDenied("Only admins and mentors can create clubs.")
        
        instance = serializer.save()
        log_action(self.request.user, f"Created Club: {instance.name}")


class ClubMemberViewSet(viewsets.ModelViewSet):
    """
    Club member management. Support adding, removing, and updating member roles.
    """
    queryset = ClubMember.objects.all()
    serializer_class = ClubMemberSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        club_id = self.request.query_params.get('club_id')
        if club_id:
            return ClubMember.objects.filter(club_id=club_id)
        return super().get_queryset()
    
    def perform_create(self, serializer):
        club_member = serializer.save()
        log_action(self.request.user, f"Added {club_member.student.user.username} to club {club_member.club.name}")


class ClubRoleViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only view for club roles and their permissions.
    """
    queryset = ClubRole.objects.all()
    serializer_class = ClubRoleSerializer
    permission_classes = [IsAuthenticated]


# ============================================================================
# EVENT MANAGEMENT
# ============================================================================

class EventViewSet(viewsets.ModelViewSet):
    queryset = Event.objects.all().select_related('aicte_category', 'club')
    serializer_class = EventSerializer
    permission_classes = [IsClubAdmin]

    def get_permissions(self):
        """
        Permissions:
        - Students: can list & retrieve events and register
        - Club Organizers/Admin: can create/update/upload attendance/generate certificates
        """
        if self.action in [
            'create', 'update', 'partial_update', 'destroy',
            'upload_attendance', 'generate_certificates',
            'start', 'end', 'mark_attendance'
        ]:
            self.permission_classes = [IsClubAdmin]  # organizers only

        elif self.action in ['register']:
            self.permission_classes = [IsStudent]  # student registration

        elif self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]  # <-- FIX: allow students

        return super().get_permissions()
    
    def get_queryset(self):
        user = self.request.user

        # students should see only scheduled events
        if hasattr(user, 'student_profile'):
            return Event.objects.filter(status='scheduled')

        # organizers/admin see all
        return Event.objects.all()

    def perform_create(self, serializer):
        user = self.request.user
        if user.user_type != 'club_organizer':
            raise PermissionDenied("Only club organizers can create events.")
        club_organizer_profile = getattr(user, "club_organizer_profile", None)
        if not club_organizer_profile or not club_organizer_profile.club:
            raise PermissionDenied("You must be assigned to a club to create events.")
        event = serializer.save(club=club_organizer_profile.club, created_by=user)
        log_action(user, f"Created Event: {event.name} (AICTE: {getattr(event.aicte_category,'name','None')}, Points: {event.points_awarded})")

    def perform_update(self, serializer):
        event = serializer.save()
        user = self.request.user

        # Check if status was changed to 'scheduled' - automatically assign hall
        if event.status == 'scheduled' and hasattr(event, '_original_status') and event._original_status != 'scheduled':
            self._assign_hall_to_event(event)

        log_action(user, f"Updated Event: {event.name}")

    def _assign_hall_to_event(self, event):
        """
        Automatically assign hall to event based on preferences.
        Creates notifications if assignment fails.
        """
        # Skip if already assigned
        if event.assigned_hall:
            return

        # Try to assign hall using event's assign_hall method
        hall_assigned = event.assign_hall()
        event.save()

        if hall_assigned:
            # Create hall booking for the assigned hall
            try:
                HallBooking.objects.create(
                    hall=event.assigned_hall,
                    event=event,
                    booking_date=event.event_date,
                    start_time=event.start_time,
                    end_time=event.end_time or event.start_time,
                    booking_status='APPROVED',
                    booked_by=event.created_by,
                    approved_by=event.created_by  # Auto-approved since preferences checked
                )
                log_action(event.created_by, f"Auto-assigned hall {event.assigned_hall.name} to event {event.name}")
            except Exception as e:
                # Log error but don't prevent event creation
                print(f"Error creating hall booking: {e}")
        else:
            # No hall available - create notification for organizer
            organizer_profile = getattr(event.club, 'organizers', None)
            if organizer_profile and organizer_profile.exists():
                organizer = organizer_profile.first().user
                Notification.objects.create(
                    user=organizer,
                    title="Hall Assignment Failed",
                    message=f"No halls available for event '{event.name}' on {event.event_date} from {event.start_time}. Please reschedule or contact admin."
                )
            log_action(event.created_by, f"Failed to assign hall to event {event.name} - no availability")


    def retrieve(self, request, *args, **kwargs):
        """Override retrieve to set _original_status for comparison in update"""
        instance = self.get_object()
        instance._original_status = instance.status
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def participants(self, request, pk=None):
        """Return participant list (id, usn, full_name) for preview prior to attendance upload"""
        event = self.get_object()
        regs = EventRegistration.objects.filter(event=event).select_related('student__user')
        results = []
        for r in regs:
            student = r.student
            results.append({
                'student_id': student.id,
                'usn': getattr(student, 'usn', ''),
                'full_name': getattr(student, 'full_name', getattr(student, 'user').username if getattr(student,'user',None) else '')
            })
        return Response({'participants': results})
    
    @action(detail=True, methods=['post'])
    def register(self, request, pk=None):
        """Students register for an event"""
        user = request.user

        if user.user_type != 'student':
            raise PermissionDenied("Only students can register for events.")

        event = self.get_object()

        # ensure event can be registered
        if event.status not in ['scheduled', 'upcoming']:
            return Response(
                {"error": "Event is not open for registration."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # fetch student profile
        try:
            student = user.student_profile
        except:
            return Response(
                {"error": "Student profile not found."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # create or ensure registration
        reg, created = EventRegistration.objects.get_or_create(
            event=event,
            student=student
        )

        if not created:
            return Response({"message": "Already registered."})

        log_action(user, f"Registered for event {event.name}")
        return Response({"message": "Registered successfully!"}, status=200)

    @action(detail=True, methods=['post'], url_path='upload-attendance')
    def upload_attendance(self, request, pk=None):
        """
        Accepts file upload (CSV or XLSX). The file should have a header with at least:
          - student_id OR usn (prefer student_id)
          - attendance (present/cancelled/no-show) or is_present (1/0/true/false)
        If no file provided and request contains 'attendance' array, uses that.
        """
        event = self.get_object()

        # Only organizers of the club that owns the event can upload attendance
        user = request.user
        if user.user_type != 'club_organizer':
            raise PermissionDenied("Only club organizers can upload attendance for events.")

        # If multipart file present
        file_obj = request.FILES.get('file')
        attendance_items = []

        if file_obj:
            filename = file_obj.name.lower()
            if filename.endswith('.csv'):
                try:
                    decoded = file_obj.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(decoded))
                    for row in reader:
                        attendance_items.append(row)
                except Exception as e:
                    return Response({'error': f'Failed to parse CSV: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                try:
                    wb = openpyxl.load_workbook(file_obj, data_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {headers[i]: row[i] for i in range(len(headers))}
                        attendance_items.append(row_dict)
                except Exception as e:
                    return Response({'error': f'Failed to parse Excel: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'error': 'Unsupported file type; only .csv and .xlsx allowed'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            # fallback to json attendance array
            attendance_items = request.data.get('attendance', [])

        if not isinstance(attendance_items, list):
            return Response({'error': 'Attendance payload must be a list'}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        points_created = 0
        certificates_created = 0

        with transaction.atomic():
            for item in attendance_items:
                # support student_id or usn; normalize keys (strings)
                student_id = item.get('student_id') or item.get('studentId') or item.get('id')
                usn = item.get('usn') or item.get('USN')
                attendance_flag = item.get('attendance') or item.get('status') or item.get('is_present')
                # normalize attendance flag
                is_present = False
                if attendance_flag is None:
                    # if missing, default to present
                    is_present = True
                else:
                    s = str(attendance_flag).strip().lower()
                    if s in ['present', 'p', '1', 'true', 'yes', 'attended']:
                        is_present = True
                    else:
                        is_present = False

                try:
                    if student_id:
                        student = Student.objects.get(id=int(student_id))
                    elif usn:
                        student = Student.objects.get(usn=usn)
                    else:
                        # skip if cannot identify
                        continue
                except Student.DoesNotExist:
                    continue

                attendance_record, created = EventAttendance.objects.update_or_create(
                    event=event,
                    student=student,
                    defaults={'is_present': is_present, 'marked_by': request.user}
                )
                if created:
                    created_count += 1

                # allocate AICTE points in PENDING if present and event awards points
                if is_present and event.aicte_category and event.points_awarded > 0:
                    AICTEPointTransaction.objects.create(
                        student=student,
                        event=event,
                        category=event.aicte_category,
                        points_allocated=event.points_awarded,
                        status='PENDING',
                        created_at=now()
                    )
                    points_created += 1

                # generate basic certificate placeholder (Certificate model may need file generation later)
                if is_present:
                    cert, cert_created = Certificate.objects.get_or_create(event=event, student=student)
                    if cert_created:
                        # create a QR placeholder - not full PDF generation here
                        qr_data = f"Certificate ID: {cert.id}, Event: {event.id}, Student: {student.usn}"
                        qr = qrcode.make(qr_data)
                        # If Certificate has an image/file field:
                        # buffer = io.BytesIO()
                        # qr.save(buffer, format='PNG')
                        # cert.qr_image.save(f"cert_{cert.id}_qr.png", ContentFile(buffer.getvalue()))
                        certificates_created += 1

            # mark event completed if not already
            if event.status != 'completed':
                event.status = 'completed'
                event.save()
            log_action(request.user, f"Uploaded attendance for event {event.name}: created {created_count} attendance records, {points_created} AICTE transactions, {certificates_created} certificates")

        return Response({
            'message': 'Attendance processed',
            'attendance_records_created': created_count,
            'aicte_transactions_created': points_created,
            'certificates_created': certificates_created
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='generate-certificates')
    def generate_certificates(self, request, pk=None):
        """Generate certificates for event participants (idempotent)"""
        from os import path
        import hashlib

        event = self.get_object()
        if event.status != 'completed':
            raise ValidationError("Certificates can only be generated for completed events.")

        attendees = EventAttendance.objects.filter(event=event, is_present=True).select_related('student__user')
        if not attendees.exists():
            raise ValidationError("No attendees found for this event.")

        # Determine template type based on AICTE category
        template_type = "certificate_regular"
        if event.aicte_category:
            template_type = "certificate_aicte"

        template_dir = path.join(settings.MEDIA_ROOT, "certificates", "templates")
        template_path = path.join(template_dir, f"{template_type}.png")
        metadata_path = path.join(template_dir, f"{template_type}.json")

        if not path.exists(template_path) or not path.exists(metadata_path):
            raise ValidationError(f"Certificate template or metadata file not found: {template_type}")

        # Get signature paths
        faculty_signature_path = ""
        principal_signature_path = ""

        # Try to get faculty signature from club coordinator
        if event.club and event.club.faculty_coordinator:
            mentor = event.club.faculty_coordinator
            if hasattr(mentor, 'signature') and mentor.signature:
                faculty_signature_path = path.join(settings.MEDIA_ROOT, str(mentor.signature))
            elif hasattr(mentor.user, 'signature') and mentor.user.signature:
                faculty_signature_path = path.join(settings.MEDIA_ROOT, str(mentor.user.signature))

        # Try to get principal signature from PrincipalSignature model
        active_principal_sig = PrincipalSignature.objects.filter(is_active=True).first()
        if active_principal_sig and active_principal_sig.signature_image:
            principal_signature_path = path.join(settings.MEDIA_ROOT, str(active_principal_sig.signature_image))

        # Delete existing certificates for this event to enable re-generation
        existing_certs = Certificate.objects.filter(event=event)
        existing_count = existing_certs.count()
        if existing_count > 0:
            for cert in existing_certs:
                # Delete the file from storage
                if cert.file:
                    cert.file.delete(save=False)
            existing_certs.delete()
            log_action(request.user, f"Deleted {existing_count} existing certificates for event: {event.name}")

        certificate_count = 0
        errors = []

        for attendance in attendees:
            student = attendance.student
            try:
                # Create new certificate record (force new since old ones deleted)
                certificate = Certificate.objects.create(
                    event=event,
                    student=student,
                    issue_date=now()
                )

                # Generate certificate using CertificateGenerator
                generator = CertificateGenerator(template_path, metadata_path)

                # Generate QR code text
                qr_text = f"Certificate ID: {certificate.id}, Student: {student.usn}, Event: {event.name}"

                # Generate certificate PDF
                pdf_buffer = generator.generate_certificate(
                    template_type=template_type,
                    student_name=student.user.get_full_name() or student.user.username,
                    event_name=event.name,
                    club_name=event.club.name if event.club else "",
                    date=event.event_date.strftime("%d %b %Y"),
                    usn=getattr(student, 'usn', 'N/A'),
                    points=event.points_awarded if event.aicte_category else 0,
                    qr_text=qr_text,
                    faculty_signature_path=faculty_signature_path,
                    principal_signature_path=principal_signature_path,
                )

                # Save PDF to certificate record with proper naming
                file_name = f"certificate_{certificate.id}.pdf"
                certificate.file.save(file_name, ContentFile(pdf_buffer.getvalue()), save=True)

                # Generate hash for verification
                pdf_buffer.seek(0)
                file_hash = hashlib.sha256(pdf_buffer.read()).hexdigest()
                certificate.file_hash = file_hash
                certificate.save()

                certificate_count += 1

                # Create notifications for students
                Notification.objects.create(
                    user=student.user,
                    title="Certificate Generated",
                    message=f"Your certificate for '{event.name}' has been generated successfully.",
                    notification_type="certificate_generated",
                    event=event,
                    certificate=certificate
                )

            except Exception as e:
                errors.append(f"Student {student.user.username}: {str(e)}")
                continue

        log_action(request.user, f"Generated {certificate_count} certificates for event: {event.name}")

        response_data = {
            'message': f'Successfully generated {certificate_count} certificates for event "{event.name}".',
            'certificate_count': certificate_count
        }

        if errors:
            response_data['errors'] = errors
            response_data['warning'] = f"{len(errors)} certificates failed to generate."

        return Response(response_data)


class EventAttendanceViewSet(viewsets.ModelViewSet):
    """
    Event attendance tracking.
    """
    queryset = EventAttendance.objects.all()
    serializer_class = EventAttendanceSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        event_id = self.request.query_params.get('event_id')
        if event_id:
            return EventAttendance.objects.filter(event_id=event_id)
        return super().get_queryset()


# ============================================================================
# HALL BOOKING MANAGEMENT
# ============================================================================

class HallViewSet(viewsets.ModelViewSet):
    """
    Hall/venue management.
    Provides endpoints for listing halls and checking availability.
    """
    queryset = Hall.objects.filter(is_available=True)
    serializer_class = HallSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def available(self, request):
        """
        Get available halls for a specific date and time range.
        Query params:
        - date: (YYYY-MM-DD) required
        - start_time: (HH:MM or HH:MM:SS) required
        - end_time: (HH:MM or HH:MM:SS) required
        """
        date_str = request.query_params.get('date')
        start_time_str = request.query_params.get('start_time')
        end_time_str = request.query_params.get('end_time')

        if not all([date_str, start_time_str, end_time_str]):
            return Response(
                {'error': 'date, start_time, and end_time query parameters are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from datetime import datetime
            booking_date = datetime.strptime(date_str, '%Y-%m-%d').date()

            # Handle both HH:MM and HH:MM:SS formats
            for time_str in [start_time_str, end_time_str]:
                if ':' not in time_str or time_str.count(':') > 2:
                    raise ValueError("Invalid time format")

            try:
                start_time = datetime.strptime(start_time_str, '%H:%M:%S').time()
            except ValueError:
                try:
                    start_time = datetime.strptime(start_time_str, '%H:%M').time()
                except ValueError:
                    raise ValueError("Invalid start_time format")

            try:
                end_time = datetime.strptime(end_time_str, '%H:%M:%S').time()
            except ValueError:
                try:
                    end_time = datetime.strptime(end_time_str, '%H:%M').time()
                except ValueError:
                    raise ValueError("Invalid end_time format")

        except ValueError as e:
            return Response(
                {'error': f'Invalid date or time format: {str(e)}. Use YYYY-MM-DD and HH:MM or HH:MM:SS'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all available halls
        all_halls = Hall.objects.filter(is_available=True)
        
        # Find halls with conflicts
        conflicting_hall_ids = HallBooking.objects.filter(
            booking_date=booking_date,
            booking_status__in=['APPROVED', 'PENDING'],
            start_time__lt=end_time,
            end_time__gt=start_time
        ).values_list('hall_id', flat=True).distinct()
        
        # Available halls are those without conflicts
        available_halls = all_halls.exclude(id__in=conflicting_hall_ids)
        
        serializer = self.get_serializer(available_halls, many=True)
        return Response(serializer.data)


class HallBookingViewSet(viewsets.ModelViewSet):
    queryset = HallBooking.objects.all().select_related('hall', 'event', 'booked_by')
    serializer_class = HallBookingSerializer
    permission_classes = [IsClubAdmin]  # creation by club organizers; admin endpoints checked below

    def get_permissions(self):
        # keep default but individual actions enforce role
        if self.action in ['approve', 'reject', 'list_admin_pending']:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()

    def _has_conflict(self, hall, booking_date, start_time, end_time):
        # returns True if any approved or pending booking overlaps
        qs = HallBooking.objects.filter(hall=hall, booking_date=booking_date).exclude(pk=self.kwargs.get('pk', None))
        # consider both pending and approved to avoid overlaps
        qs = qs.filter(booking_status__in=['PENDING', 'APPROVED'])
        for b in qs:
            # convert to comparable times
            if b.start_time <= end_time and start_time <= b.end_time:
                return True
        return False

    def create(self, request, *args, **kwargs):
        """
        Create booking: if no conflict, auto-approve (booking_status='APPROVED').
        If conflict exists, create as PENDING and leave for admin.
        """
        user = request.user
        if user.user_type != 'club_organizer':
            raise PermissionDenied("Only club organizers can create hall bookings")

        data = request.data.copy()
        hall_id = data.get('hall')
        booking_date = data.get('booking_date')
        start_time_str = data.get('start_time')
        end_time_str = data.get('end_time')

        if not all([hall_id, booking_date, start_time_str, end_time_str]):
            return Response({'error': 'hall, booking_date, start_time, end_time required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            hall = Hall.objects.get(id=hall_id)
        except Hall.DoesNotExist:
            return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

        try:
            start_time = datetime.strptime(start_time_str, "%H:%M").time()
            end_time = datetime.strptime(end_time_str, "%H:%M").time()
            booking_date_obj = datetime.strptime(booking_date, "%Y-%m-%d").date()
        except ValueError:
            return Response({'error': 'Invalid date/time format'}, status=status.HTTP_400_BAD_REQUEST)

        conflict = self._has_conflict(hall, booking_date_obj, start_time, end_time)

        # set booking_status based on conflict
        data['booking_status'] = 'PENDING' if conflict else 'APPROVED'
        data['booked_by'] = request.user.id

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            booking = serializer.save()
            if not conflict:
                booking.approved_by = request.user  # automatic approval by system (user)
                booking.save()
                log_action(request.user, f"Auto-approved hall booking: {booking.hall.name} on {booking.booking_date} {booking.start_time}-{booking.end_time}")
            else:
                log_action(request.user, f"Created pending hall booking: {booking.hall.name} on {booking.booking_date} {booking.start_time}-{booking.end_time}")

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Admin approves a pending booking"""
        if request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can approve bookings")
        booking = self.get_object()
        if booking.booking_status == 'APPROVED':
            return Response({'detail': 'Booking already approved'}, status=status.HTTP_400_BAD_REQUEST)
        # check again for conflicts before approving
        if self._has_conflict(booking.hall, booking.booking_date, booking.start_time, booking.end_time):
            return Response({'error': 'Conflicting booking found — cannot approve'}, status=status.HTTP_409_CONFLICT)
        booking.booking_status = 'APPROVED'
        booking.approved_by = request.user
        booking.save()
        log_action(request.user, f"Approved hall booking: {booking.hall.name} for {booking.event.name if booking.event else 'N/A'}")
        return Response({'message': 'Hall booking approved', 'status': booking.booking_status})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Admin rejects a pending booking with reason"""
        if request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can reject bookings")
        booking = self.get_object()
        reason = request.data.get('reason', 'No reason provided')
        if not reason or len(reason) < 5:
            return Response({'error': 'Rejection reason must be at least 5 characters'}, status=status.HTTP_400_BAD_REQUEST)
        booking.booking_status = 'REJECTED'
        booking.rejection_reason = reason
        booking.approved_by = request.user
        booking.save()
        log_action(request.user, f"Rejected hall booking: {booking.hall.name} - Reason: {reason}")
        return Response({'message': 'Hall booking rejected', 'reason': reason})

    @action(detail=False, methods=['get'])
    def list_admin_pending(self, request):
        """Admin: list pending bookings (for dashboard)"""
        if request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can view pending bookings")
        qs = HallBooking.objects.filter(booking_status='PENDING').select_related('hall', 'event', 'booked_by')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)


# ============================================================================
# CERTIFICATE MANAGEMENT
# ============================================================================

class CertificateTemplateViewSet(viewsets.ModelViewSet):
    """
    Global certificate template management.
    Only administrators can create/update templates.
    """
    queryset = CertificateTemplate.objects.all()
    serializer_class = CertificateTemplateSerializer
    permission_classes = [IsAuthenticated]
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            self.permission_classes = [IsAuthenticated]
        return super().get_permissions()
    
    def perform_create(self, serializer):
        if self.request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can create certificate templates.")
        
        template = serializer.save(created_by=self.request.user)
        log_action(self.request.user, f"Created certificate template: {template.name}")
    
    def perform_update(self, serializer):
        if self.request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can update certificate templates.")
        
        template = serializer.save()
        log_action(self.request.user, f"Updated certificate template: {template.name}")


class CertificateViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Certificate viewing. Students can view their own, mentors can view mentees'.
    """
    serializer_class = CertificateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'student':
            return Certificate.objects.filter(student__user=user)
        elif user.user_type == 'mentor':
            mentees = Student.objects.filter(mentor__user=user)
            return Certificate.objects.filter(student__in=mentees)
        elif user.user_type == 'admin':
            return Certificate.objects.all()
        return Certificate.objects.none()

    @action(detail=False, methods=['get'], url_path='verify/(?P<file_hash>[^/.]+)')
    def verify(self, request, file_hash=None):
        """Verify certificate by file hash"""
        cert = Certificate.objects.filter(file_hash=file_hash).select_related('student', 'event').first()
        
        if not cert:
            return Response(
                {'verified': False, 'message': 'Certificate not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        return Response({
            'verified': True,
            'certificate_id': cert.id,
            'student_usn': cert.student.usn,
            'event_name': cert.event.name,
            'issue_date': cert.issue_date,
        })


# ============================================================================
# AICTE POINTS MANAGEMENT
# ============================================================================

class AICTECategoryViewSet(viewsets.ModelViewSet):
    """
    AICTE point categories and rules.
    Only administrators can create/modify.
    """
    queryset = AICTECategory.objects.all()
    serializer_class = AICTECategorySerializer
    permission_classes = [IsAuthenticated]
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            if self.request.user.user_type != 'admin':
                raise PermissionDenied("Only administrators can manage AICTE categories.")
        return super().get_permissions()


class AICTEPointTransactionViewSet(viewsets.ModelViewSet):
    """
    AICTE point transactions with mentor approval workflow.
    """
    queryset = AICTEPointTransaction.objects.all()
    serializer_class = AICTEPointTransactionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.user_type == 'student':
            return AICTEPointTransaction.objects.filter(student__user=user)
        elif user.user_type == 'mentor':
            mentees = Student.objects.filter(mentor__user=user)
            return AICTEPointTransaction.objects.filter(student__in=mentees)
        elif user.user_type == 'admin':
            return AICTEPointTransaction.objects.all()
        return AICTEPointTransaction.objects.none()

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Mentor approval of AICTE points"""
        tx = self.get_object()
        mentor = getattr(request.user, 'mentor_profile', None)
        
        if not mentor or tx.student.mentor != mentor:
            raise PermissionDenied("Only assigned mentor can approve points.")
        
        tx.status = 'APPROVED'
        tx.approved_by = request.user
        tx.approval_date = now()
        tx.save()
        log_action(request.user, f"Approved AICTE transaction ID {tx.id}")
        
        return Response({'message': 'Points approved successfully.'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Mentor rejection of AICTE points"""
        tx = self.get_object()
        mentor = getattr(request.user, 'mentor_profile', None)
        
        if not mentor or tx.student.mentor != mentor:
            raise PermissionDenied("Only assigned mentor can reject points.")
        
        tx.status = 'REJECTED'
        tx.rejection_reason = request.data.get('reason', '')
        tx.approved_by = request.user
        tx.approval_date = now()
        tx.save()
        log_action(request.user, f"Rejected AICTE transaction ID {tx.id}")
        
        return Response({'message': 'Points rejected.'})


# ============================================================================
# NOTIFICATION MANAGEMENT
# ============================================================================

class NotificationViewSet(viewsets.ModelViewSet):
    """
    User notifications (email, in-app, SMS).
    """
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user).order_by('-created_at')

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        """Mark all notifications as read"""
        notifications = self.get_queryset()
        notifications.update(is_read=True)
        log_action(request.user, "Marked all notifications as read")

        return Response({'message': 'All notifications marked as read.'})

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """Mark single notification as read"""
        notification = self.get_object()
        notification.is_read = True
        notification.save()

        return Response({'message': 'Notification marked as read.'})

    @action(detail=True, methods=['post'])
    def take_action(self, request, pk=None):
        """Take action on notification based on its type"""
        notification = self.get_object()
        action = request.data.get('action')
        user = request.user

        if notification.notification_type == 'certificate_generated' and notification.event:
            # Open certificate download
            certificate = Certificate.objects.filter(
                event=notification.event,
                student__user=user
            ).first()
            if certificate:
                return Response({
                    'type': 'download',
                    'certificate_id': certificate.id,
                    'file_url': request.build_absolute_uri(certificate.file.url) if certificate.file else None
                })
            return Response({'error': 'Certificate not found'}, status=404)

        elif notification.notification_type == 'points_approved' and notification.aicte_transaction:
            # Open AICTE points verification page
            return Response({
                'type': 'redirect',
                'url': f'/dashboard?section=aicte_points',
                'transaction_id': notification.aicte_transaction.id
            })

        elif notification.notification_type == 'points_rejected' and notification.aicte_transaction:
            # Open AICTE points verification page
            return Response({
                'type': 'redirect',
                'url': f'/dashboard?section=aicte_points',
                'transaction_id': notification.aicte_transaction.id
            })

        elif notification.notification_type == 'hall_booking_approved' and notification.hall_booking:
            # Open hall booking details
            return Response({
                'type': 'redirect',
                'url': f'/dashboard?section=hall_bookings',
                'booking_id': notification.hall_booking.id
            })

        elif notification.notification_type == 'hall_booking_rejected' and notification.hall_booking:
            # Open hall booking details
            return Response({
                'type': 'redirect',
                'url': f'/dashboard?section=hall_bookings',
                'booking_id': notification.hall_booking.id
            })

        elif notification.notification_type == 'event_registration':
            # Open event details
            if notification.event:
                return Response({
                    'type': 'redirect',
                    'url': f'/dashboard?section=events&event_id={notification.event.id}'
                })

        elif notification.notification_type == 'event_reminder':
            # Open event details
            if notification.event:
                return Response({
                    'type': 'redirect',
                    'url': f'/dashboard?section=events&event_id={notification.event.id}'
                })

        # Mark as read after taking action
        notification.is_read = True
        notification.save()

        log_action(user, f"Took action on notification {notification.id}: {action}")
        return Response({'message': 'Action completed.'})


# ============================================================================
# AUDIT & REPORTING
# ============================================================================

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Audit logs - admin only access.
    """
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.user_type != 'admin':
            raise PermissionDenied("Only administrators can view audit logs.")
        return super().get_queryset().order_by('-timestamp')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    """Get dashboard statistics based on user role"""
    user = request.user
    
    if user.user_type == 'student':
        student = getattr(user, 'student_profile', None)
        if not student:
            return Response({'error': 'Student profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({
            'role': 'student',
            'total_aicte_points': student.total_aicte_points,
            'events_registered': student.event_registrations.count(),
            'certificates_earned': Certificate.objects.filter(student=student).count(),
            'pending_approvals': AICTEPointTransaction.objects.filter(
                student=student,
                status='PENDING'
            ).count(),
        })
    
    elif user.user_type == 'mentor':
        mentor = getattr(user, 'mentor_profile', None)
        if not mentor:
            return Response({'error': 'Mentor profile not found'}, status=status.HTTP_404_NOT_FOUND)
        
        mentees = Student.objects.filter(mentor=mentor)
        pending_approvals = AICTEPointTransaction.objects.filter(
            student__in=mentees,
            status='PENDING'
        ).count()
        
        return Response({
            'role': 'mentor',
            'total_mentees': mentees.count(),
            'pending_approvals': pending_approvals,
            'profile_completed': mentor.profile_completed,
        })
    
    elif user.user_type == 'admin':
        return Response({
            'role': 'admin',
            'total_users': User.objects.count(),
            'total_students': Student.objects.count(),
            'total_mentors': Mentor.objects.count(),
            'total_clubs': Club.objects.count(),
            'total_events': Event.objects.count(),
        })
    
    return Response({'error': 'Invalid user type'}, status=status.HTTP_400_BAD_REQUEST)
